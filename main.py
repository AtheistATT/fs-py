from openpyxl import *
import os
import datetime
import pdb

data = {}
total = []


def convert_to_delta(s):
    minutes, seconds = s.split(':')
    seconds, milliseconds = seconds.split('.')
    return datetime.timedelta(minutes=int(minutes), seconds=int(seconds), microseconds=int(milliseconds))

def convert_to_string(t):
    total = int(t.total_seconds())
    minutes, seconds = divmod(total, 60)
    microseconds = int(t.microseconds / 10000)
    return f"{minutes:02}:{seconds:02}.{microseconds:02}"


def check_file():
    if not os.path.isfile("input.xlsx"):
        wb = Workbook()
        wb.remove(wb["Sheet"])

        for i in range(10):
            wb.create_sheet(f"Эстафета{i + 1}")
            for _ in range(5):
                wb[f"Эстафета{i + 1}"].append(["Школа","00:00.00"])
        wb.save("input.xlsx")
        wb.close()
        exit()

def load_data():
    wb = load_workbook("input.xlsx")

    for sheet in wb.sheetnames:
        if sheet not in data.keys():
            data[sheet] = []
        for row in wb[sheet].iter_rows(values_only=True):
            data[sheet].append(list(row))


def sort_data():
    for key in data.keys():
        for t in data[key]:
            t[1] = convert_to_delta(t[1])
        data[key] = sorted(data[key], key=lambda x:x[1])

        data[key] = [[i + 1, data[key][i][0], data[key][i][1]] for i in range(len(data[key]))]


def get_total():

    total_dict = {}
    for key in data.keys():
        for row in data[key]:
            if row[1] not in total_dict.keys():
                total_dict[row[1]] = 0
            total_dict[row[1]] += row[0]

    global total
    total = [[key, val] for key, val in total_dict.items()]
    total = sorted(total, key=lambda x:x[1])

    total = [[i+1, total[i][1], total[i][0]] for i in range(len(total))]

    for i in range(1, len(total)):
        if total[i][1] == total[i - 1][1]:
            total[i][0] = total[i - 1][0]

def save_data():
    wb = Workbook()

    wb.remove(wb["Sheet"])

    for key in data.keys():
        wb.create_sheet(key)
        for row in data[key]:
            row[2] = convert_to_string(row[2])
            wb[key].append(row)

    wb.create_sheet("TOTAL")

    for row in total:
        wb["TOTAL"].append(row)

    wb.save("output.xlsx")
    wb.close()



check_file()
load_data()
sort_data()
get_total()
save_data()
